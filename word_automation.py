from selenium import webdriver
import openpyxl
import pandas as pd
import time
import subprocess
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

EXCEL = 'words.xlsx'
driver = webdriver.Chrome('chromedriver.exe')

def search_english_meaning(word):
    driver.get("https://www.ldoceonline.com/jp/")
    time.sleep(1)
    driver.find_element_by_xpath('/html/body/div[2]/form/div[2]/input').send_keys(word)
    driver.find_element_by_xpath('/html/body/div[2]/form/button').click()
    try:
        ans = driver.find_element_by_css_selector('span.Head + span.Sense > span.DEF').text
    except:
        ans = 'no answer'
    return ans

def search_english_synonyms(word):
    ans = []
    driver.get('https://www.thesaurus.com/')
    time.sleep(1)
    driver.find_element_by_xpath('/html/body/div/div/div/div[1]/header/div/div/div/form/div[2]/input').send_keys(word)
    driver.find_element_by_xpath('/html/body/div/div/div/div[1]/header/div/div/div/form/button').click()
    time.sleep(1)
    for i in range(4):
        try:
            ans.append(driver.find_element_by_css_selector(f"#initial-load-content > main > section > section > div:nth-child(2) > section:nth-child(1) > ul > li:nth-child({i+1}) > span > a").text)
        except:
            ans.append('no answer')
    return [a for a in ans if a != '']

'''
print(search_english_synonyms('primary'))
driver.close()
driver.quit()
'''

df = pd.read_excel(EXCEL)
words = df.values.tolist()
meanings = []
synonyms = []

for word in words:
    meanings.append(search_english_meaning(word[0]))

for word in words:
    synonyms.append(search_english_synonyms(word[0]))

driver.close()
driver.quit()

wb = openpyxl.load_workbook(EXCEL)

sheet = wb['Sheet1']

for i in range(len(meanings)):
    sheet.cell(row=i+2, column=2, value=meanings[i])
    for j in range(len(synonyms[i])):
        sheet.cell(row=i+2, column=3+j, value=synonyms[i][j])
        
wb.save(EXCEL)
subprocess.Popen(['start',EXCEL], shell=True)

